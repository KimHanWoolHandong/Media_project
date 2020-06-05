from konlpy.tag import Komoran
import csv, codecs
import openpyxl


"""
komoran = Komoran()
content = "내가 지금 무엇을 하고 있을까?"
content_morphs = komoran.pos(content)
print(content_morphs)
"""
news_lst = ['경향신문', '국민일보', '동아일보', '문화일보', '서울신문', '세계일보', '조선일보', '중앙일보',
            '한겨레', '한국일보', '매일경제', '한국경제']
month_to_quater = {'01':'1', '02':'1', '03':'1', '04':'2', '05':'2', '06':'2', '07':'3', '08':'3', '09':'3',
                   '10':'4', '11':'4', '12':'4'}
def calculate_csv(csv_locate, pos_dic, neg_dic, komoran):
    news = {}
    count = {}
    with codecs.open(csv_locate, 'r', encoding='utf-8') as csvf:
        rd = csv.reader(csvf)
        next(rd)
        i = 0
        for line in rd:
            if len(line) <= 1:
                pass
            elif line[1] not in news_lst:
                pass
            else:
                news_year = line[0][:-4]
                news_month = line[0][4:6]
                date = news_year+month_to_quater[news_month]
                #print(date)
                news_n_date = line[1]+"/"+date
                morphs = komoran.pos(line[2])
                score = cal_pos_neg(morphs, pos_dic, neg_dic)
                if news_n_date in news.keys():
                    news[news_n_date] = score + news[news_n_date]
                    #print(line[1] + " " + str(score))
                    count[news_n_date] = 1 + count[news_n_date]
                else:
                    news[news_n_date] = score
                    count[news_n_date] = 1
            #print(line[0])
            i += 1
    #print(news)
    #print(count)
    average = {}
    for name in news.keys():
        if count[name] == 0:
            average[name] = 'None'
        else:
            average[name] = news[name] / count[name]
    #print(average)
    print(news)
    return news, count, average

def dictionary_creating():
    pos_dic = {}
    with codecs.open('./pos.csv', 'r', encoding='utf-8') as csvf:
        rd = csv.reader(csvf)
        for line in rd:
            char = line[0].split('+')
            text_comprehensive = []
            for text in char:
                morph_n_type = text.split('/')
                text_comprehensive.append((morph_n_type[0],morph_n_type[1]))
            if len(text_comprehensive) == 1:
                pos_dic[(text_comprehensive[0])] = line[1]
            else:
                pos_dic[tuple(text_comprehensive)] = line[1]

    neg_dic = {}
    with codecs.open('./neg.csv', 'r', encoding='utf-8') as csvf:
        rd = csv.reader(csvf)
        for line in rd:
            char = line[0].split('+')
            #print(char)
            text_comprehensive = []
            for text in char:
                morph_n_type = text.split('/')
                text_comprehensive.append((morph_n_type[0], morph_n_type[1]))
            if len(text_comprehensive) == 1:
                neg_dic[(text_comprehensive[0])] = line[1]
            else:
                neg_dic[tuple(text_comprehensive)] = line[1]

    return pos_dic, neg_dic

def find_word(word, pos_dic, neg_dic):
    Flag = True
    pos = 0
    neg = 0
    if word in pos_dic.keys():
        #print('pos')
        pos = float(pos_dic[word])
        Flag = False
    if word in neg_dic.keys():
        #print('neg')
        neg = float(neg_dic[word])
        Flag = False
    return pos, neg, Flag

def cal_pos_neg(content, pos_dic, neg_dic):
    #print(content)
    length = len(content)
    neg_count = 0
    pos_count = 0
    for i in range(length):
        if length-i >= 3:
            three_word = (content[i], content[i+1], content[i+2])
            pos, neg, Flag = find_word(three_word, pos_dic, neg_dic)
            if Flag:
                second_word = (content[i], content[i + 1])
                pos, neg, Flag = find_word(second_word, pos_dic, neg_dic)
                if Flag:
                    word = (content[i])
                    pos, neg, Flag = find_word(word, pos_dic, neg_dic)
                    pos_count += pos
                    neg_count += neg
                else:
                    pos_count += pos
                    neg_count += neg
            else:
                pos_count += pos
                neg_count += neg
        elif length-i == 2:
            second_word = (content[i], content[i + 1])
            pos, neg, Flag = find_word(second_word, pos_dic, neg_dic)
            if Flag:
                word = (content[i])
                pos, neg, Flag = find_word(word, pos_dic, neg_dic)
                pos_count += pos
                neg_count += neg
            else:
                pos_count += pos
                neg_count += neg
        else:
            word = (content[i])
            pos, neg, Flag = find_word(word, pos_dic, neg_dic)
            pos_count += pos
            neg_count += neg
    #print(pos_count)
    #print(neg_count)
    score = 0
    if pos_count > neg_count:
        score = pos_count / (pos_count + neg_count)
    elif neg_count > pos_count:
        score = -neg_count / (pos_count + neg_count)
    else:
        score = 0
    return score

def write_csv(president, news, count, average):
    xl = openpyxl.load_workbook('./quater_result.xlsx')
    sheet = xl.get_sheet_by_name(president)
    i = 1
    for key, data in news.items():
        news_n_date = key.split('/')
        news_name = news_n_date[0]
        date = news_n_date[1]

        sheet.cell(row=i, column=1).value = news_name
        sheet.cell(row=i, column=2).value = date
        sheet.cell(row=i, column=3).value = data
        sheet.cell(row=i, column=4).value = count[key]
        sheet.cell(row=i, column=5).value = average[key]
        i += 1
    xl.save("./quater_result.xlsx")
    xl.close()





def calculation_collection(pos_dic, neg_dic, komoran):
    donga_news, donga_count, donga_average = calculate_csv('./donga.csv', pos_dic, neg_dic, komoran)
    josun_news, josun_count, josun_average = calculate_csv('./josun.csv', pos_dic, neg_dic, komoran)

    write_csv('Donga', donga_news, donga_count, donga_average)
    write_csv('Josun', josun_news, josun_count, josun_average)


komoran = Komoran()
pos_dic, neg_dic = dictionary_creating()
calculation_collection(pos_dic, neg_dic, komoran)


"""
content = "지금 상황이 안좋네요"

content_morphs = komoran.pos(content)
cal_pos_neg(content_morphs, pos_dic, neg_dic)
"""





